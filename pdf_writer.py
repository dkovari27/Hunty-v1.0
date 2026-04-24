"""
pdf_writer.py — Generates a clean PDF job report from Excel output or job list.

Portrait A4, three columns: Job Title · Location · Link (clickable).
Only NEW jobs are included by default (new_only=True).
"""
from __future__ import annotations

import logging
import os
from datetime import datetime

import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

logger = logging.getLogger(__name__)

# ── Layout ─────────────────────────────────────────────────────────────────────
_PAGE_W, _PAGE_H = A4        # 595 × 842 pt, portrait
_MARGIN          = 18 * mm   # left/right/top/bottom

_CONTENT_W = _PAGE_W - 2 * _MARGIN   # ~559 pt

# Column widths (must sum to _CONTENT_W)
_W_TITLE    = _CONTENT_W * 0.52   # Job Title  — widest, wraps freely
_W_LOCATION = _CONTENT_W * 0.26   # Location
_W_LINK     = _CONTENT_W * 0.22   # Clickable link

# ── Colours ────────────────────────────────────────────────────────────────────
_DARK_BLUE  = colors.HexColor("#1F4E79")
_MID_BLUE   = colors.HexColor("#2E75B6")
_LINK_BLUE  = colors.HexColor("#0563C1")
_STRIPE     = colors.HexColor("#EEF4FA")
_BORDER     = colors.HexColor("#C8D8E8")
_GREY_TEXT  = colors.HexColor("#555555")


# ── Stats table colours ────────────────────────────────────────────────────────
_STAT_SECTION = colors.HexColor("#D6E4F0")
_STAT_NEW     = colors.HexColor("#D4EDDA")
_STAT_STRIPE  = colors.HexColor("#F5F9FC")
_NEW_TEXT     = colors.HexColor("#155724")


# ── Paragraph styles ───────────────────────────────────────────────────────────
def _make_styles() -> tuple:
    title_style = ParagraphStyle(
        "report_title",
        fontName="Helvetica-Bold",
        fontSize=15,
        textColor=_DARK_BLUE,
        alignment=1,   # centred
        spaceAfter=6,
    )
    header_style = ParagraphStyle(
        "col_header",
        fontName="Helvetica-Bold",
        fontSize=9,
        textColor=colors.white,
        leading=13,
    )
    cell_style = ParagraphStyle(
        "cell_body",
        fontName="Helvetica",
        fontSize=8.5,
        leading=12,
        textColor=colors.black,
    )
    link_style = ParagraphStyle(
        "cell_link",
        fontName="Helvetica",
        fontSize=8.5,
        leading=12,
        textColor=_LINK_BLUE,
    )
    return title_style, header_style, cell_style, link_style


# ── Run-summary table ──────────────────────────────────────────────────────────

def _build_stats_table(stats: dict) -> Table:
    """Centred two-column table summarising scrape stats for the PDF header."""
    _FONT  = "Helvetica"
    _FSIZE = 8.5
    _PAD   = 8

    hdr  = ParagraphStyle("sh",  fontName="Helvetica-Bold",        fontSize=_FSIZE, textColor=colors.white, leading=12)
    hdrc = ParagraphStyle("shc", fontName="Helvetica-Bold",        fontSize=_FSIZE, textColor=colors.white, leading=12, alignment=1)
    sec  = ParagraphStyle("ss",  fontName="Helvetica-BoldOblique", fontSize=8,      textColor=_DARK_BLUE,   leading=11, alignment=1)
    lbl  = ParagraphStyle("sl",  fontName=_FONT,                   fontSize=_FSIZE, textColor=colors.black, leading=12)
    lbl_c= ParagraphStyle("slc", fontName=_FONT,                   fontSize=_FSIZE, textColor=colors.black, leading=12, alignment=1)
    val  = ParagraphStyle("sv",  fontName=_FONT,                   fontSize=_FSIZE, textColor=colors.black, leading=12, alignment=1)
    nlbl = ParagraphStyle("snl", fontName="Helvetica-Bold",        fontSize=_FSIZE, textColor=_NEW_TEXT,    leading=12, alignment=1)
    nval = ParagraphStyle("snv", fontName="Helvetica-Bold",        fontSize=_FSIZE, textColor=_NEW_TEXT,    leading=12, alignment=1)
    _e   = Paragraph("", lbl)

    sources  = stats.get("sources", [])
    pipeline = [
        ("    Total scraped",            stats.get("total_scraped", 0)),
        ("    Duplicates removed",        stats.get("duplicates_removed", 0)),
        ("    Unique jobs",               stats.get("unique_jobs", 0)),
        ("    Dropped by keyword filter", stats.get("dropped_by_filter", 0)),
        ("    Passed to export",          stats.get("passed_to_export", 0)),
    ]

    # ── Column width: longest label + 6 char-widths + padding ────────────────
    fixed_labels = [
        "  Sources", "  Pipeline",
        "    Total scraped", "    Duplicates removed", "    Unique jobs",
        "    Dropped by keyword filter", "    Passed to export",
        "New vs previous run", "Total run time", "Metric",
    ]
    source_labels = [f"    {name}" for name, _, _ in sources]
    max_text_w = max(stringWidth(s, _FONT, _FSIZE) for s in fixed_labels + source_labels)
    col_w = max_text_w + stringWidth("x" * 6, _FONT, _FSIZE) + _PAD * 2
    cnt_w = col_w / 2

    rows       = []
    style_cmds = []

    def _bg(c0, c1, ri, colour):
        style_cmds.append(("BACKGROUND", (c0, ri), (c1, ri), colour))

    def _span(c0, c1, ri):
        style_cmds.append(("SPAN", (c0, ri), (c1, ri)))

    # ── Row 0: column headers ─────────────────────────────────────────────────
    rows.append([Paragraph("Metric", hdr), Paragraph("Count", hdrc),
                 Paragraph("Metric", hdr), Paragraph("Count", hdrc)])
    _bg(0, 3, 0, _DARK_BLUE)

    # ── Row 1: section labels side by side ───────────────────────────────────
    rows.append([Paragraph("  Sources", sec),  _e,
                 Paragraph("  Pipeline", sec), _e])
    _bg(0, 1, 1, _STAT_SECTION); _span(0, 1, 1)
    _bg(2, 3, 1, _STAT_SECTION); _span(2, 3, 1)

    # ── Data rows: sources left, pipeline right (pad shorter side) ────────────
    n = max(len(sources), len(pipeline))
    for i in range(n):
        left_lbl = left_val = _e
        if i < len(sources):
            sname, scount, _ = sources[i]
            left_lbl = Paragraph(f"    {sname}", lbl)
            left_val = Paragraph(str(scount), val)

        right_lbl = right_val = _e
        if i < len(pipeline):
            plabel, pcount = pipeline[i]
            right_lbl = Paragraph(plabel, lbl)
            right_val = Paragraph(str(pcount), val)

        rows.append([left_lbl, left_val, right_lbl, right_val])
        if i % 2 == 0:
            _bg(0, 3, len(rows) - 1, _STAT_STRIPE)

    # ── New vs previous run — spans full left half / right half ──────────────
    new_idx = len(rows)
    rows.append([Paragraph("New vs previous run", nlbl), _e,
                 Paragraph(str(stats.get("new_count", 0)), nval), _e])
    _bg(0, 3, new_idx, _STAT_NEW)
    _span(0, 1, new_idx); _span(2, 3, new_idx)

    # ── Total run time ────────────────────────────────────────────────────────
    run_s    = stats.get("run_time_s", 0)
    run_str  = f"{int(run_s // 60)}m {int(run_s % 60)}s" if run_s >= 60 else f"{int(run_s)}s"
    time_idx = len(rows)
    rows.append([Paragraph("Total run time", lbl_c), _e,
                 Paragraph(run_str, val), _e])
    _span(0, 1, time_idx); _span(2, 3, time_idx)

    # ── Shared structural styles ──────────────────────────────────────────────
    style_cmds += [
        ("BOX",          (0, 0), (-1, -1), 0.8, _MID_BLUE),
        ("LINEBELOW",    (0, 0), (-1, -1), 0.4, _BORDER),
        ("LINEAFTER",    (0, 0), (-1, -1), 0.4, _BORDER),
        ("LINEAFTER",    (1, 0), (1, -1),  1.2, _MID_BLUE),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
        ("LEFTPADDING",  (0, 0), (-1, -1), _PAD),
        ("RIGHTPADDING", (0, 0), (-1, -1), _PAD),
    ]

    tbl = Table(rows, colWidths=[col_w, cnt_w, col_w, cnt_w], hAlign="CENTER")
    tbl.setStyle(TableStyle(style_cmds))
    return tbl


# ── Excel loader ───────────────────────────────────────────────────────────────

def load_jobs_from_excel(
    excel_path: str,
    new_only: bool = True,
    max_jobs: int | None = None,
) -> list[dict]:
    """
    Read jobs from a Hunty Excel output file.
    Returns list of dicts with keys: title, location, url, _status.
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb.active

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}

    jobs: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        status = str(row[col.get("Status", 0)] or "").strip()
        if new_only and status != "NEW":
            continue
        jobs.append({
            "title":    str(row[col["Job Title"]] or "").strip(),
            "location": str(row[col["Location"]]  or "").strip(),
            "url":      str(row[col["URL"]]        or "").strip(),
            "_status":  status,
        })
        if max_jobs and len(jobs) >= max_jobs:
            break

    wb.close()
    logger.info("Loaded %d jobs from %s (new_only=%s)", len(jobs), excel_path, new_only)
    return jobs


# ── PDF builder ────────────────────────────────────────────────────────────────

def generate_pdf(
    jobs: list[dict],
    output_path: str,
    report_title: str = "Hunty — New Job Listings",
    stats: dict | None = None,
) -> str:
    """
    Write a PDF job report to output_path.

    Each job dict must have: title, location, url.
    stats — optional run-summary dict from run_job_scraper(); when provided
            a summary table is inserted between the header and jobs table.
    Returns output_path.
    """
    title_style, header_style, cell_style, link_style = _make_styles()

    # ── Header row ────────────────────────────────────────────────────────────
    header_row = [
        Paragraph("Job Title",  header_style),
        Paragraph("Location",   header_style),
        Paragraph("Link",       header_style),
    ]

    # ── Data rows ─────────────────────────────────────────────────────────────
    rows: list[list] = [header_row]
    for job in jobs:
        url = job.get("url", "") or ""
        if url:
            # Show as much of the URL as fits in the column (~30 chars), full URL as target
            display = url if len(url) <= 20 else url[:18] + "…"
            link_cell = Paragraph(
                f'<a href="{url}"><font color="#0563C1"><u>{display}</u></font></a>',
                link_style,
            )
        else:
            link_cell = Paragraph("—", cell_style)

        rows.append([
            Paragraph(job.get("title", "") or "—", cell_style),
            Paragraph(job.get("location", "") or "—", cell_style),
            link_cell,
        ])

    # ── Table style ───────────────────────────────────────────────────────────
    style_cmds = [
        # Header background
        ("BACKGROUND",   (0, 0), (-1, 0), _DARK_BLUE),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _STRIPE]),
        # Grid
        ("BOX",          (0, 0), (-1, -1), 0.8, _MID_BLUE),
        ("LINEBELOW",    (0, 0), (-1, -1), 0.4, _BORDER),
        ("LINEAFTER",    (0, 0), (-1, -1), 0.4, _BORDER),
        # Alignment & padding
        ("VALIGN",       (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING",   (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
        ("LEFTPADDING",  (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        # Link column: centre vertically
        ("VALIGN",       (2, 1), (2, -1), "MIDDLE"),
    ]

    table = Table(
        rows,
        colWidths=[_W_TITLE, _W_LOCATION, _W_LINK],
        repeatRows=1,
        hAlign="LEFT",
    )
    table.setStyle(TableStyle(style_cmds))

    # ── Document ──────────────────────────────────────────────────────────────
    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        leftMargin=_MARGIN,
        rightMargin=_MARGIN,
        topMargin=_MARGIN,
        bottomMargin=_MARGIN,
        title=report_title,
    )

    date_label = datetime.now().strftime("%d/%m/%y")

    story: list = [
        Paragraph(f"Hunty — {date_label}", title_style),
        Spacer(1, 6),
    ]
    if stats:
        story.append(_build_stats_table(stats))
        story.append(Spacer(1, 10))
    story.append(table)

    def _footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(_GREY_TEXT)
        canvas.drawCentredString(
            _PAGE_W / 2,
            12 * mm,
            f"Page {doc.page}",
        )
        canvas.restoreState()

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    logger.info("PDF saved: %s (%d jobs)", output_path, len(jobs))
    return output_path


# ── CLI test ───────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import glob
    import sys

    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    # Find latest Excel in outputs/
    pattern = os.path.join("outputs", "jobs_*.xlsx")
    files = sorted(glob.glob(pattern))
    if not files:
        print("No Excel output found in outputs/")
        sys.exit(1)

    latest = files[-1]
    print(f"Source: {latest}")

    # Load first 10 jobs regardless of new/seen status (test mode)
    jobs = load_jobs_from_excel(latest, new_only=False, max_jobs=10)
    print(f"Loaded {len(jobs)} jobs")

    out = os.path.join("outputs", "test_report.pdf")
    generate_pdf(jobs, out, report_title="Hunty — Test Report (10 jobs)")
    print(f"PDF written: {out}")
