"""
excel_writer.py — Writes filtered job listings to a formatted Excel workbook.

Color coding:
  Green  (score 8–10) — strong match
  Yellow (score 6–7)  — good match
  Red    (score < 6)  — weak match
"""
import json
import logging
import math
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Distance from Basel — haversine lookup
# ---------------------------------------------------------------------------

_BASEL_LAT, _BASEL_LON = 47.5596, 7.5886

# Approximate coordinates for European pharma hubs and common job locations.
# None = remote / no meaningful distance.
_CITY_COORDS: dict[str, tuple[float, float] | None] = {
    # Switzerland
    "basel": (47.5596, 7.5886),
    "allschwil": (47.5498, 7.5333),
    "binningen": (47.5370, 7.5699),
    "birsfelden": (47.5476, 7.6236),
    "muttenz": (47.5226, 7.6468),
    "pratteln": (47.5147, 7.6911),
    "reinach": (47.4958, 7.5953),
    "arlesheim": (47.4966, 7.6206),
    "kaiseraugst": (47.5374, 7.7627),
    "sisseln": (47.5522, 7.8669),
    "riehen": (47.5786, 7.6411),
    "zurich": (47.3769, 8.5417),
    "zürich": (47.3769, 8.5417),
    "schlieren": (47.3978, 8.4458),
    "kloten": (47.4504, 8.5818),
    "bern": (46.9481, 7.4474),
    "berne": (46.9481, 7.4474),
    "zug": (47.1661, 8.5152),
    "baar": (47.1957, 8.5294),
    "rotkreuz": (47.1425, 8.4281),
    "lausanne": (46.5197, 6.6323),
    "geneva": (46.2044, 6.1432),
    "genf": (46.2044, 6.1432),
    "genève": (46.2044, 6.1432),
    "lucerne": (47.0502, 8.3093),
    "luzern": (47.0502, 8.3093),
    "visp": (46.2942, 7.8812),
    "winterthur": (47.5001, 8.7248),
    "st. gallen": (47.4245, 9.3767),
    "st gallen": (47.4245, 9.3767),
    "olten": (47.3520, 7.9041),
    "aarau": (47.3910, 8.0432),
    "solothurn": (47.2088, 7.5352),
    "neuchatel": (46.9900, 6.9293),
    "neuchâtel": (46.9900, 6.9293),
    "fribourg": (46.8065, 7.1619),
    "schaffhausen": (47.6973, 8.6345),
    "thun": (46.7580, 7.6275),
    "sion": (46.2330, 7.3582),
    "wädenswil": (47.2283, 8.6748),
    # Germany
    "munich": (48.1351, 11.5820),
    "münchen": (48.1351, 11.5820),
    "frankfurt": (50.1109, 8.6821),
    "berlin": (52.5200, 13.4050),
    "hamburg": (53.5753, 10.0153),
    "freiburg": (47.9990, 7.8421),
    "freiburg im breisgau": (47.9990, 7.8421),
    "heidelberg": (49.3988, 8.6724),
    "mannheim": (49.4875, 8.4660),
    "düsseldorf": (51.2217, 6.7762),
    "dusseldorf": (51.2217, 6.7762),
    "cologne": (50.9333, 6.9500),
    "köln": (50.9333, 6.9500),
    "koln": (50.9333, 6.9500),
    "stuttgart": (48.7758, 9.1829),
    "darmstadt": (49.8728, 8.6512),
    "mainz": (49.9929, 8.2473),
    "biberach": (48.0971, 9.7927),
    "ingelheim": (49.9677, 8.0713),
    "leverkusen": (51.0459, 6.9841),
    "marburg": (50.8021, 8.7747),
    "wiesbaden": (50.0820, 8.2417),
    # UK
    "london": (51.5074, -0.1278),
    "cambridge": (52.2053, 0.1218),
    "oxford": (51.7520, -1.2577),
    "stevenage": (51.9020, -0.2024),
    "macclesfield": (53.2583, -2.1284),
    "manchester": (53.4808, -2.2426),
    "edinburgh": (55.9533, -3.1883),
    "glasgow": (55.8642, -4.2518),
    "birmingham": (52.4862, -1.8904),
    "leeds": (53.8008, -1.5491),
    "bristol": (51.4545, -2.5879),
    "slough": (51.5105, -0.5950),
    "guildford": (51.2362, -0.5704),
    # France
    "paris": (48.8566, 2.3522),
    "strasbourg": (48.5734, 7.7521),
    "lyon": (45.7640, 4.8357),
    "marseille": (43.2965, 5.3698),
    "suresnes": (48.8721, 2.2252),
    # Netherlands
    "amsterdam": (52.3676, 4.9041),
    "leiden": (52.1601, 4.4970),
    "utrecht": (52.0907, 5.1214),
    "rotterdam": (51.9244, 4.4777),
    "eindhoven": (51.4416, 5.4697),
    "oss": (51.7653, 5.5196),
    # Belgium
    "brussels": (50.8503, 4.3517),
    "bruxelles": (50.8503, 4.3517),
    "ghent": (51.0500, 3.7167),
    "gent": (51.0500, 3.7167),
    "leuven": (50.8798, 4.7005),
    # Denmark
    "copenhagen": (55.6761, 12.5683),
    "bagsvaerd": (55.7500, 12.4500),
    # Sweden
    "stockholm": (59.3293, 18.0686),
    "gothenburg": (57.7089, 11.9746),
    "göteborg": (57.7089, 11.9746),
    "malmo": (55.6050, 13.0038),
    "malmö": (55.6050, 13.0038),
    "lund": (55.7047, 13.1910),
    "molndal": (57.6561, 12.0114),
    "mölndal": (57.6561, 12.0114),
    # Austria
    "vienna": (48.2082, 16.3738),
    "wien": (48.2082, 16.3738),
    "graz": (47.0707, 15.4395),
    # Italy
    "milan": (45.4642, 9.1900),
    "milano": (45.4642, 9.1900),
    "rome": (41.9028, 12.4964),
    # Spain
    "barcelona": (41.3851, 2.1734),
    "madrid": (40.4168, -3.7038),
    # Ireland
    "dublin": (53.3498, -6.2603),
    # Norway
    "oslo": (59.9139, 10.7522),
    # Finland
    "helsinki": (60.1699, 24.9384),
    # USA
    "boston": (42.3601, -71.0589),
    "new york": (40.7128, -74.0060),
    "san francisco": (37.7749, -122.4194),
    "san diego": (32.7157, -117.1611),
    "chicago": (41.8781, -87.6298),
    "seattle": (47.6062, -122.3321),
    "los angeles": (34.0522, -118.2437),
    "new jersey": (40.0583, -74.4057),
    "north chicago": (42.3253, -87.8412),
    "south san francisco": (37.6547, -122.4077),
    "indianapolis": (39.7684, -86.1581),
    "tarrytown": (41.0759, -73.8596),
    "durham": (35.9940, -78.8986),
    "emeryville": (37.8310, -122.2858),
    "gaithersburg": (39.1434, -77.2014),
    "thousand oaks": (34.1706, -118.8376),
    # Remote / null
    "remote": None,
    "home office": None,
    "hybrid": None,
    "worldwide": None,
    "anywhere": None,
}


def _haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> int:
    R = 6371
    d_lat = math.radians(lat2 - lat1)
    d_lon = math.radians(lon2 - lon1)
    a = (math.sin(d_lat / 2) ** 2
         + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2))
         * math.sin(d_lon / 2) ** 2)
    return round(2 * R * math.asin(math.sqrt(a)))


def dist_from_basel(location: str) -> int | None:
    """
    Return approximate straight-line distance from Basel in km.
    Returns None if location is unknown, empty, or remote.
    """
    if not location:
        return None
    loc = location.lower().strip()
    # Try exact match on full string, then on each comma-separated token
    candidates = [loc] + [p.strip() for p in loc.split(",")]
    for c in candidates:
        if c in _CITY_COORDS:
            coords = _CITY_COORDS[c]
            return None if coords is None else _haversine_km(_BASEL_LAT, _BASEL_LON, *coords)
    # Substring match (only keys longer than 4 chars to avoid "bern" in "bernhard" etc.)
    for key, coords in _CITY_COORDS.items():
        if len(key) > 4 and key in loc:
            return None if coords is None else _haversine_km(_BASEL_LAT, _BASEL_LON, *coords)
    return None


# ---------------------------------------------------------------------------
# Persistent seen-job history
# ---------------------------------------------------------------------------

_HISTORY_FILENAME = "seen_jobs_history.json"


def load_previous_urls(output_dir: str) -> dict[str, str]:
    """
    Return {job_url: first_seen_date_iso} for all previously seen jobs.
    Backed by seen_jobs_history.json in output_dir.
    """
    path = os.path.join(output_dir, _HISTORY_FILENAME)
    if os.path.isfile(path):
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                history = {str(k): str(v) for k, v in data.items()}
                logger.info("History: %d known URLs loaded", len(history))
                return history
        except Exception as exc:
            logger.warning("Could not read job history: %s", exc)
    return {}


def _save_history(output_dir: str, history: dict[str, str]) -> None:
    """Persist the seen-job URL → first_seen_date mapping."""
    os.makedirs(output_dir, exist_ok=True)
    path = os.path.join(output_dir, _HISTORY_FILENAME)
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(history, f, indent=2, ensure_ascii=False)
    except Exception as exc:
        logger.warning("Could not save job history: %s", exc)


# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

# Application-status colour fills (applied to the App. Status cell)
_APP_STATUS_FILLS = {
    "Saved":       PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "Applied":     PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "Interviewing":PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid"),
    "Offer":       PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Rejected":    PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    "Accepted":    PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),
    "Withdrawn":   PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
}

# (header label, job dict key, column width)
_COLUMNS = [
    ("Status",            "_status",           8),
    ("App. Status",       "_app_status",       13),
    ("Score",             "relevance_score",   8),
    ("Job Title",         "title",             36),
    ("Company",           "company",           24),
    ("Location",          "location",          20),
    ("Dist. Basel (km)",  "distance_basel_km", 16),
    ("Date Posted",       "date_posted",       14),
    ("First Seen",        "first_seen",        11),
    ("URL",               "url",               52),
    ("Source",            "source",            12),
    ("AI Reasoning",      "ai_reasoning",      56),
]

_RAW_COLUMNS = [
    ("Job Title",         "title",             36),
    ("Company",           "company",           24),
    ("Location",          "location",          20),
    ("Dist. Basel (km)",  "distance_basel_km", 16),
    ("Date Posted",       "date_posted",       14),
    ("URL",               "url",               52),
    ("Source",            "source",            12),
]


def _thin_border() -> Border:
    side = Side(style="thin", color="D0D0D0")
    return Border(left=side, right=side, top=side, bottom=side)


def _score_fill(score: int) -> PatternFill:
    if score == 0:
        return PatternFill(fill_type=None)   # no fill — unscored job
    if score >= 8:
        color = "C6EFCE"   # green
    elif score >= 6:
        color = "FFEB9C"   # yellow
    else:
        color = "FFC7CE"   # red
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _enrich_distance(job: dict) -> dict:
    """Add distance_basel_km to a job dict (non-destructive copy)."""
    d = dict(job)
    if "distance_basel_km" not in d:
        d["distance_basel_km"] = dist_from_basel(job.get("location", ""))
    return d


def _write_urls_sheet(wb: openpyxl.Workbook, jobs: list[dict]) -> None:
    """Sheet 2 — one clickable URL row per scraped job, sorted by source then title."""
    ws = wb.create_sheet("All URLs")

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border       = _thin_border()

    cols = [
        ("Source",   "source",   14),
        ("Job Title","title",    40),
        ("Company",  "company",  26),
        ("Location", "location", 22),
        ("URL",      "url",      60),
    ]
    for col_i, (label, _, width) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_i, value=label)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_i)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    sorted_jobs = sorted(jobs, key=lambda j: (j.get("source", ""), j.get("title", "").lower()))
    for row_i, job in enumerate(sorted_jobs, start=2):
        ws.row_dimensions[row_i].height = 16
        for col_i, (_, key, _) in enumerate(cols, start=1):
            value = job.get(key, "")
            cell = ws.cell(row=row_i, column=col_i, value=value)
            cell.border    = border
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            if key == "url" and value:
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")

    ws.auto_filter.ref = ws.dimensions


def write_excel(jobs: list[dict], all_jobs: list[dict] | None = None,
                filename: str | None = None,
                previous_urls: dict[str, str] | None = None,
                applications: dict[str, dict] | None = None) -> tuple[str, int]:
    """
    Write jobs to an Excel file in OUTPUT_DIR.
    Returns (full_path, new_job_count).

    previous_urls:  {url: first_seen_date} from load_previous_urls().
    applications:   {url: {status, title, company, updated}} from application_tracker.load().
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(OUTPUT_DIR, filename or f"jobs_{timestamp}.xlsx")

    # Build a mutable copy of history so we can add new entries
    history: dict[str, str] = dict(previous_urls) if previous_urls else {}
    today_str = datetime.now().strftime("%Y-%m-%d")

    jobs = [_enrich_distance(j) for j in jobs]
    for job in jobs:
        url = job.get("url", "").strip()
        if not url or url not in history:
            job["_status"] = "NEW"
            job["first_seen"] = ""   # blank; will show as "NEW" already
            if url:
                history[url] = today_str  # record first appearance
        else:
            job["_status"] = "—"
            job["first_seen"] = history[url]

    # Persist updated history so the next run knows about today's jobs
    _save_history(OUTPUT_DIR, history)

    # Inject application status from tracker
    apps = applications or {}
    for job in jobs:
        url = job.get("url", "").strip()
        job["_app_status"] = apps.get(url, {}).get("status", "")

    new_count = sum(1 for j in jobs if j["_status"] == "NEW")

    # Sort: new jobs first, then previously seen; within each group by distance / score
    jobs.sort(key=lambda j: (
        0 if j["_status"] == "NEW" else 1,
        j["distance_basel_km"] if isinstance(j.get("distance_basel_km"), (int, float)) else 99999,
        -j.get("relevance_score", 0),
    ))

    wb = openpyxl.Workbook()

    # Fills for the Status column
    _new_fill  = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # light blue
    _seen_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # light gray

    # ------------------------------------------------------------------ #
    # Sheet 1 — Job Results
    # ------------------------------------------------------------------ #
    ws = wb.active
    ws.title = "Job Results"

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border       = _thin_border()

    for col_i, (label, _, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_i, value=label)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_i)].width = width

    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    for row_i, job in enumerate(jobs, start=2):
        score    = job.get("relevance_score", 0)
        is_new   = job["_status"] == "NEW"
        row_fill = _score_fill(score) if is_new else _seen_fill
        ws.row_dimensions[row_i].height = 58

        for col_i, (_, key, _) in enumerate(_COLUMNS, start=1):
            value = job.get(key, "")
            if isinstance(value, list):
                value = " • ".join(value) if value else ""

            cell = ws.cell(row=row_i, column=col_i, value=value)
            cell.border    = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if key == "_status":
                cell.fill      = _new_fill if is_new else _seen_fill
                cell.font      = Font(bold=True, color="1F4E79" if is_new else "999999", size=9)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif key == "_app_status":
                app_fill = _APP_STATUS_FILLS.get(value)
                if app_fill:
                    cell.fill = app_fill
                cell.font      = Font(bold=bool(value), size=9,
                                      color="155724" if value == "Accepted" else "000000")
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=False)
            elif not is_new:
                # De-emphasise seen rows: gray text, light gray background
                cell.fill = _seen_fill
                cell.font = Font(color="888888")
                if key == "url" and value:
                    cell.hyperlink = value
                    cell.font = Font(color="888888", underline="single")
                elif key == "relevance_score":
                    cell.font      = Font(bold=True, size=12, color="AAAAAA")
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                elif key == "distance_basel_km":
                    cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.fill = row_fill
                if key == "url" and value:
                    cell.hyperlink = value
                    cell.font = Font(color="0563C1", underline="single")
                elif key == "relevance_score":
                    cell.font      = Font(bold=True, size=12)
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                elif key == "distance_basel_km":
                    cell.alignment = Alignment(horizontal="center", vertical="top")

    ws.auto_filter.ref = ws.dimensions

    # ------------------------------------------------------------------ #
    # Sheet 2 — All URLs (every scraped job as a clickable link)
    # ------------------------------------------------------------------ #
    url_source = all_jobs if all_jobs is not None else jobs
    _write_urls_sheet(wb, url_source)

    # ------------------------------------------------------------------ #
    # Sheet 3 — Summary
    # ------------------------------------------------------------------ #
    ws3 = wb.create_sheet("Summary")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 24

    ws3["A1"] = "Hunty — Run Summary"
    ws3["A1"].font = Font(bold=True, size=14, color="1F4E79")

    scored = [j for j in jobs if j.get("relevance_score", 0) > 0]
    rows = [
        ("Run date",              datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total scraped (raw)",   len(all_jobs) if all_jobs is not None else "—"),
        ("Passed pre-filter",     len(jobs)),
        ("New jobs (not seen before)", new_count),
        ("Previously seen jobs",  len(jobs) - new_count),
    ]
    if scored:
        scores = [j["relevance_score"] for j in scored]
        rows += [
            ("Scored by AI",          len(scored)),
            ("Strong matches (8-10)", sum(1 for s in scores if s >= 8)),
            ("Good matches (6-7)",    sum(1 for s in scores if 6 <= s < 8)),
            ("Below threshold (< 6)", sum(1 for s in scores if s < 6)),
            ("Highest score",         max(scores)),
            ("Average score",         round(sum(scores) / len(scores), 1)),
        ]
    else:
        rows.append(("AI scoring", "disabled"))

    for r_i, (label, value) in enumerate(rows, start=3):
        ws3.cell(row=r_i, column=1, value=label).font = Font(bold=True)
        ws3.cell(row=r_i, column=2, value=value)

    wb.save(filepath)
    logger.info(
        "Excel saved: %s  (%d jobs, %d new, %d raw)",
        filepath, len(jobs), new_count, len(all_jobs) if all_jobs else 0,
    )
    return filepath, new_count


def export_jobs_json(jobs: list[dict], output_path: str) -> None:
    """
    Write jobs to a JSON side-output file.

    Wraps the list in {"jobs": [...], "count": N, "timestamp": "...", "source": "scraper_export"}.
    Non-serialisable values (datetime objects, numpy types, etc.) are coerced to str.
    """
    from datetime import datetime

    def _safe(v):
        if isinstance(v, (str, int, float, bool, list, dict)) or v is None:
            return v
        return str(v)

    payload = {
        "jobs":      [{k: _safe(v) for k, v in job.items()} for job in jobs],
        "count":     len(jobs),
        "timestamp": datetime.now().isoformat(),
        "source":    "scraper_export",
    }

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
    logger.info("JSON export: %s (%d jobs)", output_path, len(jobs))
