"""Build swiss_companies.xlsx from collected research data."""
import os
import sys
from collections import Counter
from pathlib import Path
from urllib.parse import urlparse

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

os.makedirs("output", exist_ok=True)

# Load _OVERRIDES from the registry to classify scraper config per company URL
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
try:
    from scrapers.european._registry import _OVERRIDES as _REG_OVERRIDES
except ImportError:
    _REG_OVERRIDES = {}


def _scraper_config_label(url: str) -> str:
    """Return a human-readable scraper-config label for a company career-page URL."""
    parsed = urlparse(url)
    domain = parsed.netloc
    dom_path = domain + parsed.path.rstrip("/")
    overrides = _REG_OVERRIDES.get(dom_path) or _REG_OVERRIDES.get(domain)
    if overrides is None:
        return "Auto – listing (base URL)"
    scraper = overrides.get("scraper", "generic")
    if scraper == "skip":
        return "Manual – check direct"
    search_url = overrides.get("search_url", "")
    if "{keyword}" in search_url:
        return "Auto – keyword search"
    if search_url:
        return "Auto – listing page"
    return "Auto – listing (base URL)"

companies = [
    # (Company, City, Type, Career Page URL, Notes)
    # ── BASEL CITY ────────────────────────────────────────────────────────────
    ("Boehringer Ingelheim",    "Basel",     "Pharma",  "https://jobs.boehringer-ingelheim.com/",                        "Major pharma; NBE-Therapeutics acquired"),
    ("Regeneron",               "Basel",     "Pharma",  "https://careers.regeneron.com/en/",                             "Grosspeter Tower Basel office"),
    ("Roche",                   "Basel",     "Pharma",  "https://careers.roche.com/global/en",                          ""),
    ("Novartis",                "Basel",     "Pharma",  "https://www.novartis.com/careers",                             ""),
    ("Sandoz",                  "Basel",     "Pharma",  "https://www.sandoz.com/careers/",                              "Novartis spin-off, generics"),
    ("Sanofi",                  "Basel",     "Pharma",  "https://jobs.sanofi.com/en",                                   "CH offices Basel & Vernier"),
    ("Bayer",                   "Basel",     "Pharma",  "https://www.bayer.com/en/ch/careers",                          "Consumer Health & Pharma hub Basel"),
    ("BeOne Medicines",         "Basel",     "Pharma",  "https://beigene.wd3.myworkdayjobs.com/beigene",                "Formerly BeiGene, European HQ"),
    ("Moderna",                 "Basel",     "Pharma",  "https://www.modernatx.com/careers",                            "European mRNA operations hub"),
    ("Mepha",                   "Basel",     "Pharma",  "https://www.mepha.ch/de-ch/unternehmen/karriere/",             "Generics, Teva subsidiary"),
    ("Pentapharm",              "Basel",     "Pharma",  "https://www.pentapharm.com/careers",                           "Specialty biochemicals"),
    ("Swedish Orphan Biovitrum","Basel",     "Pharma",  "https://sobi.csod.com/ats/careersite/search.aspx?site=2&c=sobi", "SOBI Basel office"),
    ("Lonza",                   "Basel",     "CDMO",    "https://www.lonza.com/careers",                                "Major CDMO; multiple CH sites"),
    ("Celonic",                 "Basel",     "CDMO",    "https://celonic-ag.jobs.personio.com",                         "Biologics CDMO"),
    ("LYO-X",                   "Basel",     "CDMO",    "https://lyo-x.com/en/",                                        "Lyophilisation specialist"),
    ("ten23 health",            "Basel",     "CDMO",    "https://ten23.health/careers/",                                "Drug development & manufacture"),
    ("SpiroChem",               "Basel",     "CRO",     "https://spirochem.com/careers/",                               "Discovery chemistry, libraries, ADC/PROTAC"),
    ("Bright Peak Therapeutics","Basel",     "Biotech",  "https://brightpeaktx.com/careers/",                           "Engineered cytokines"),
    ("BioVersys",               "Basel",     "Biotech",  "https://www.bioversys.com/careers/",                          "Antibiotics"),
    ("Noema Pharma",            "Basel",     "Biotech",  "https://jobs.ashbyhq.com/noema",                              "CNS"),
    ("Monte Rosa Therapeutics", "Basel",     "Biotech",  "https://jobs.ashbyhq.com/MonteRosaTherapeutics",              "Molecular glue degraders"),
    ("NBE-Therapeutics",        "Basel",     "Biotech",  "https://nbe-therapeutics.com/employment/vacancies/",          "ADC; Boehringer Ingelheim subsidiary"),
    ("Vir Biotechnology",       "Basel",     "Biotech",  "https://jobs.ashbyhq.com/vir",                                "Infectious disease / immunology"),
    ("Roivant Sciences",        "Basel",     "Biotech",  "https://jobs.ashbyhq.com/roivant",                            ""),
    ("FoRx Therapeutics",       "Basel",     "Biotech",  "https://www.forxtherapeutics.com/career",                     ""),
    ("Cimeio Therapeutics",     "Basel",     "Biotech",  "https://www.cimeio.com/careers/",                             "Cell & gene therapy"),
    ("Mosanna Therapeutics",    "Basel",     "Biotech",  "https://mosanna.com/careers/",                                ""),
    ("Synendos Therapeutics",   "Basel",     "Biotech",  "https://www.synendos.com/open-position",                      "Endocannabinoid system"),
    ("Tolremo Therapeutics",    "Basel",     "Biotech",  "https://www.tolremo.com/careers",                             ""),
    ("Vaximm",                  "Basel",     "Biotech",  "https://vaximm.com/about-us/#Careers",                        "Oral cancer vaccines"),
    ("Versameb",                "Basel",     "Biotech",  "https://versameb.com/contact/career/",                        "mRNA therapeutics"),
    ("Genedata",                "Basel",     "Biotech",  "https://jobs.danaher.com/global/en/genedata",                  "Bioinformatics software; Danaher subsidiary"),
    ("Scailyte",                "Basel",     "Biotech",  "https://scailyte.com/careers/",                               "Single-cell analytics AI"),
    ("Oxford BioTherapeutics",  "Basel",     "Biotech",  "https://www.oxfordbiotherapeutics.com/careers",               "ADC oncology"),
    ("Skyhawk Therapeutics",    "Basel",     "Biotech",  "https://job-boards.greenhouse.io/skyhawktherapeutics",        "RNA splicing small molecules; Basel labs at Superlab"),
    ("Alentis Therapeutics",    "Basel",     "Biotech",  "https://jobs.ashbyhq.com/alentis",                            "Fibrotic diseases & cancer; $181M Series D 2024"),
    ("Windward Bio",            "Basel",     "Biotech",  "https://jobs.ashbyhq.com/WindwardBio",                        "Immunology; respiratory & dermatology; $200M Series A 2025"),
    ("Abbmira Therapeutics",    "Basel",     "Biotech",  "https://www.linkedin.com/company/abbmira/",                   "Small molecules; fine-tuning immunity; founded 2024"),
    ("Anaveon",                 "Basel",     "Biotech",  "https://anaveon.com/careers/",                                "IL-2 immunotherapy; clinical stage"),
    ("RhyGaze",                 "Basel",     "Biotech",  "https://rhygaze.com/careers/",                                "Gene therapy for retinal disease; Basel & Philadelphia"),
    ("Captor Therapeutics",     "Basel",     "Biotech",  "https://jobs.ashbyhq.com/captor",                             "Molecular glue degraders; targeted protein degradation"),
    ("Altamira Therapeutics",   "Basel",     "Biotech",  "https://www.altamiratherapeutics.com/careers/",               "RNA delivery platforms"),
    ("Alloy Therapeutics",      "Basel",     "Biotech",  "https://www.alloytx.com/careers/",                            "Antibody discovery platform; collaborative biotech"),
    ("Amporin Pharmaceuticals", "Basel",     "Biotech",  "https://amporin.com/careers/",                                "Cell membrane-protecting small molecules; founded 2024"),
    ("AlloCyte Pharmaceuticals","Basel",     "Biotech",  "https://allocyte.com/careers/",                               "Allogeneic cell therapy"),
    ("Ridgeline Discovery",     "Basel",     "CRO",     "https://ridgelinediscovery.com/careers/",                      "Drug discovery CRO; Tech Park Basel"),
    ("WuXi XDC",               "Basel",     "CDMO",    "https://www.wuxixdc.com/careers/",                             "ADC CDMO; WuXi AppTec subsidiary"),
    ("T3 Pharmaceuticals",      "Basel",     "Biotech",  "https://www.t3pharma.com/careers/",                           "Engineered bacteria targeting solid tumors"),
    ("PharmaBiome",             "Basel",     "Biotech",  "https://pharmabio.me/careers/",                               "Microbiome-based biotherapeutics"),
    ("Granite Bio",             "Basel",     "Biotech",  "https://www.granitebio.com/careers/",                         "Immunology; Tech Park Basel"),
    ("Stromal Therapeutics",    "Basel",     "Biotech",  "https://stromaltx.com/careers/",                              "Targeting tumour stroma; Tech Park Basel"),
    ("Straumann Group",         "Basel",     "MedTech", "https://www.straumann.com/group/en/discover/career.html",      "Dental implants"),
    # ── BASEL SUBURBS ─────────────────────────────────────────────────────────
    ("J&J Innovative Medicine", "Allschwil", "Pharma",  "https://www.careers.jnj.com/en/jobs/",                         "Formerly Actelion; Janssen R&D"),
    ("Idorsia Pharmaceuticals", "Allschwil", "Pharma",  "https://careers.idorsia.com/",                                 "Spun off from Actelion 2017"),
    ("Basilea Pharmaceutica",   "Allschwil", "Pharma",  "https://jobs.basilea.com/eng",                                 "Anti-infectives"),
    ("Spexis",                  "Allschwil", "Pharma",  "https://spexisbio.com/careers/",                               "Formerly Polyphor; rare disease/oncology"),
    ("Swiss Pharma Contract",   "Allschwil", "CRO",     "https://swisspharmaco.com/contact/",                           "Full-service CRO/CDMO"),
    ("CARBOGEN AMCIS",          "Bubendorf", "CDMO",    "https://www.carbogen-amcis.com/careers/open-positions",        "4 Swiss sites; API & HPAPI development"),
    ("Bachem",                  "Bubendorf", "CDMO",    "https://careers.bachem.com/search",                           "Peptide & oligonucleotide CDMO"),
    ("Siegfried",               "Zofingen",  "CDMO",    "https://siegfried.wd103.myworkdayjobs.com/external",           "~30 min from Basel; API & dosage forms"),
    # ── ZURICH CITY ───────────────────────────────────────────────────────────
    ("Takeda",                  "Zurich",    "Pharma",  "https://jobs.takeda.com/search-jobs",                         "Swiss HQ Zurich; manufacturing Neuchatel"),
    ("AbbVie",                  "Zurich",    "Pharma",  "https://careers.abbvie.com/",                                  ""),
    ("Acino",                   "Zurich",    "Pharma",  "https://acino.swiss/careers/",                                 "Specialty pharma"),
    ("Fosun Pharma",            "Zurich",    "Pharma",  "https://www.fosunpharma.com/en/careers/",                      ""),
    ("Alvotech",                "Zurich",    "Pharma",  "https://www.alvotech.com/about-us/careers",                    "Biosimilars"),
    ("MetrioPharm",             "Zurich",    "Pharma",  "https://www.metriopharm.com/en/about-us/Career.html",          "Anti-inflammatory"),
    ("CDR-Life",                "Zurich",    "Biotech",  "https://www.cdr-life.com/careers/#open-positions",            "Bispecific antibodies / T-cell engagers"),
    ("Muvon Therapeutics",      "Zurich",    "Biotech",  "https://www.muvon-therapeutics.com/openpositions",            "Cell therapy"),
    ("Cellvie",                 "Zurich",    "Biotech",  "https://cellvie.bio/careers",                                 "Mitochondrial medicine"),
    ("Endogena Therapeutics",   "Zurich",    "Biotech",  "https://endogena.com/careers/",                               "Retinal regeneration"),
    ("Mabylon",                 "Zurich",    "Biotech",  "https://mabylon.com/careers/",                                "Human antibodies; allergy & neuro"),
    ("SYDRA AG",                "Zurich",    "Biotech",  "https://sydra.io/careers/",                                   "AI-driven longevity drug discovery"),
    ("DataHow",                 "Zurich",    "Biotech",  "https://datahow.ch/about/careers/",                           "Bioprocess AI / digital twin"),
    ("QPillars",                "Zurich",    "Biotech",  "https://qpillars.com/careers",                                "Pharma quality software"),
    ("MaxWell Biosystems",      "Zurich",    "Biotech",  "https://www.mxwbio.com/careers",                              "High-density MEA electrophysiology"),
    ("BC Platforms",            "Zurich",    "Biotech",  "https://careers.bcplatforms.com/",                            "Genomic data platform"),
    ("Sequana Medical",         "Zurich",    "MedTech", "https://www.sequanamedical.com/open-positions/",               "Fluid management implants"),
    ("Celerion",                "Zurich",    "CRO",     "https://www.celerion.com/careers",                             "Clinical phase I/II CRO"),
    # ── ZURICH SUBURBS ────────────────────────────────────────────────────────
    ("Molecular Partners",      "Schlieren", "Biotech",  "https://www.molecularpartners.com/careers/",                  "DARPin therapeutics; Zurich suburb"),
    ("Philochem",               "Otelfingen","Biotech",  "https://www.philogen.com/en/careers",                         "Small-molecule drug conjugates; Philogen R&D arm"),
    ("Tandem Therapeutics",     "Zurich",    "Biotech",  "https://jobs.ashbyhq.com/tandem",                             "Extracellular matrix/fibrosis oncology; Wyss Zurich; W.A. de Vigier Award 2024"),
    # ── ZUG / ROTKREUZ ────────────────────────────────────────────────────────
    # Note: Zug has ~230 life science companies, mostly European HQs / commercial
    # offices for US biotechs. R&D-active sites most relevant for chemistry roles.
    ("Amgen",                   "Rotkreuz",  "Pharma",  "https://careers.amgen.com/en/Switzerland",                     "Full R&D & affiliate hub; Rotkreuz ZG"),
    ("Roche Diagnostics",       "Rotkreuz",  "Pharma",  "https://careers.roche.com/global/en/rotkreuz",                 "Separate diagnostics site from Basel HQ; 2700+ employees"),
    ("AstraZeneca",             "Zug",       "Pharma",  "https://careers.astrazeneca.com/location/switzerland-jobs/7684/2658434/2", "285+ employees CH; oncology/rare disease focus"),
    ("Biogen",                  "Baar",      "Biotech",  "https://www.biogen.com/careers.html",                         "International HQ; neurology; manufacturing in Luterbach"),
    ("Bristol Myers Squibb",    "Steinhausen","Pharma",  "https://careers.bms.com/",                                    "1150 employees CH; sites in Steinhausen, Zofingen, Boudry"),
    ("Gilead Sciences",         "Zug",       "Pharma",  "https://jobs.gilead.com/",                                     "European affiliate; antiviral/oncology/inflammation"),
    ("GSK",                     "Zug",       "Pharma",  "https://www.gsk.com/en-gb/careers/",                           "Oncology unit in Zug; global pharma"),
    ("Pfizer",                  "Zurich",    "Pharma",  "https://www.pfizer.com/people/careers",                        "Swiss affiliate; Zurich/Zug area office"),
    ("Galderma",                "Zug",       "Pharma",  "https://www.galderma.com/en/careers",                          "Dermatology; HQ in Zug; IPO 2024"),
    ("Blueprint Medicines",     "Zug",       "Biotech",  "https://job-boards.greenhouse.io/blueprintmedicines",         "Precision oncology; kinase/allosteric inhibitors"),
    ("Alnylam",                 "Zug",       "Biotech",  "https://www.alnylam.com/careers",                             "RNAi therapeutics; European HQ Zug"),
    ("Nuvation Bio",            "Zug",       "Biotech",  "https://nuvationbio.com/careers/",                            "Oncology; drug discovery focus"),
    ("Apellis",                 "Zug",       "Biotech",  "https://apellis.com/careers/",                                "Complement system; geographic atrophy"),
    ("Immunocore",              "Zug",       "Biotech",  "https://jobs.ashbyhq.com/immunocore",                         "ImmTAX bispecifics; T-cell engagers"),
    ("PSI CRO",                 "Zug",       "CRO",     "https://psicro.com/careers/",                                  "Global full-service CRO; Swiss-grown"),
    ("PolyPeptide Group",       "Zug",       "CDMO",    "https://www.polypeptide.com/careers",                          "Peptide & oligonucleotide CDMO"),
    ("Recipharm",               "Zug",       "CDMO",    "https://www.recipharm.com/join-us/",                           "Drug product CDMO; Zug office"),
    # ── BERN ──────────────────────────────────────────────────────────────────
    ("CSL Behring",             "Bern",      "Pharma",  "https://www.csl.com/careers",                                  "Plasma-derived therapies; large Bern site"),
    ("AlveoliX",                "Bern",      "Biotech",  "https://www.alveolix.com/about/#careers",                    "Lung-on-chip organ models"),
    ("CellnTec",                "Bern",      "Biotech",  "https://cellntec.com/careers/",                               "Epithelial cell culture systems"),
    ("Ypsomed",                 "Burgdorf",  "MedTech", "https://www.ypsomed.com/en/careers.html",                      "Drug delivery devices; 15 min from Bern"),
    # ── LAUSANNE / VAUD ───────────────────────────────────────────────────────
    ("Ferring Pharmaceuticals", "St-Prex",   "Pharma",  "https://www.ferring.com/careers/",                             "CH HQ near Lausanne; reproductive health"),
    ("Debiopharm",              "Lausanne",  "Pharma",  "https://apply.workable.com/debiopharm/",                       "Oncology & anti-infectives"),
    ("AC Immune",               "Lausanne",  "Biotech",  "https://www.acimmune.com/careers/",                           "Neurodegeneration; Alzheimer"),
    ("HAYA Therapeutics",       "Lausanne",  "Biotech",  "https://www.hayatx.com/careers",                              "lncRNA; rare disease"),
    ("Oculis",                  "Lausanne",  "Biotech",  "https://oculis.com/our-company/our-culture/#join-us",         "Eye disease"),
    ("AB2 Bio",                 "Lausanne",  "Biotech",  "http://www.ab2bio.com/en/careers.32.html",                    "IL-18 pathway; inflammatory disease"),
    ("Ichnos Sciences",         "Lausanne",  "Biotech",  "https://www.ichnossciences.com/careers/",                     "Bispecific antibodies; ex-Glenmark"),
    ("Opna Bio",                "Epalinges", "Biotech",  "https://opnabio.com/careers/",                                "Oncology; near Lausanne"),
    ("Onward Therapeutics",     "Epalinges", "Biotech",  "https://www.onwardtherapeutics.com/careers/",                 "Oncology; near Lausanne"),
    ("RS Research",             "Lausanne",  "CRO",     "https://rsresearch.net/careers/",                              "Clinical CRO"),
    ("Novostia",                "Lausanne",  "MedTech", "https://www.novostia.com/careers",                             "Cardiac surgery devices"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Swiss Companies"

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="1F4E79")

city_fills = {
    "Basel":     PatternFill("solid", fgColor="EBF3FB"),
    "Allschwil": PatternFill("solid", fgColor="EBF3FB"),
    "Bubendorf": PatternFill("solid", fgColor="EBF3FB"),
    "Zofingen":  PatternFill("solid", fgColor="EBF3FB"),
    "Zurich":     PatternFill("solid", fgColor="E9F5E9"),
    "Schlieren":  PatternFill("solid", fgColor="E9F5E9"),
    "Otelfingen": PatternFill("solid", fgColor="E9F5E9"),
    "Zug":        PatternFill("solid", fgColor="E8F8F0"),
    "Rotkreuz":   PatternFill("solid", fgColor="E8F8F0"),
    "Baar":       PatternFill("solid", fgColor="E8F8F0"),
    "Steinhausen":PatternFill("solid", fgColor="E8F8F0"),
    "Bern":      PatternFill("solid", fgColor="FFF8E7"),
    "Burgdorf":  PatternFill("solid", fgColor="FFF8E7"),
    "Lausanne":  PatternFill("solid", fgColor="FDF0F8"),
    "St-Prex":   PatternFill("solid", fgColor="FDF0F8"),
    "Epalinges": PatternFill("solid", fgColor="FDF0F8"),
}
default_fill = PatternFill("solid", fgColor="FFFFFF")
thin = Side(style="thin", color="CCCCCC")
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
url_font = Font(color="0563C1", underline="single")

headers    = ["Company", "City", "Type", "Career Page URL", "Notes", "Scraper Config"]
col_widths = [35, 14, 10, 65, 45, 22]

for c, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = brd
    ws.column_dimensions[get_column_letter(c)].width = w
ws.row_dimensions[1].height = 22

_config_fills = {
    "Auto – keyword search":  PatternFill("solid", fgColor="C6EFCE"),
    "Auto – listing page":    PatternFill("solid", fgColor="FFEB9C"),
    "Auto – listing (base URL)": PatternFill("solid", fgColor="FFEB9C"),
    "Manual – check direct":  PatternFill("solid", fgColor="FFC7CE"),
}
_config_fonts = {
    "Auto – keyword search":  Font(bold=True, color="155724"),
    "Auto – listing page":    Font(color="856404"),
    "Auto – listing (base URL)": Font(color="856404"),
    "Manual – check direct":  Font(bold=True, color="721C24"),
}

for r, (company, city, ctype, url, notes) in enumerate(companies, 2):
    row_fill     = city_fills.get(city, default_fill)
    cfg_label    = _scraper_config_label(url)
    values       = [company, city, ctype, url, notes, cfg_label]
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border    = brd
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if c == 6:
            cell.fill = _config_fills.get(cfg_label, default_fill)
            cell.font = _config_fonts.get(cfg_label, Font())
        else:
            cell.fill = row_fill
            if c == 4 and val:
                cell.font = url_font
    ws.row_dimensions[r].height = 16

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:F{len(companies) + 1}"

# Summary sheet
ws2 = wb.create_sheet("Summary")
ws2["A1"] = "Swiss Life Sciences — Career Pages"
ws2["A1"].font = Font(bold=True, size=14)
cities = Counter(c[1] for c in companies)
types  = Counter(c[2] for c in companies)
ws2["A3"] = f"Total companies: {len(companies)}"
ws2["A4"] = f"Basel area (incl. Allschwil, Bubendorf, Zofingen): {cities['Basel']+cities['Allschwil']+cities['Bubendorf']+cities['Zofingen']}"
ws2["A5"] = f"Zurich area (incl. Schlieren, Otelfingen): {cities['Zurich']+cities['Schlieren']+cities['Otelfingen']}"
ws2["A5"] = f"Zurich area (incl. Schlieren, Otelfingen): {cities['Zurich']+cities['Schlieren']+cities['Otelfingen']}"
ws2["A6"] = f"Zug area (incl. Rotkreuz, Baar, Steinhausen): {cities['Zug']+cities['Rotkreuz']+cities['Baar']+cities['Steinhausen']}"
ws2["A7"] = f"Bern area (incl. Burgdorf): {cities['Bern']+cities['Burgdorf']}"
ws2["A8"] = f"Lausanne area (incl. St-Prex, Epalinges): {cities['Lausanne']+cities['St-Prex']+cities['Epalinges']}"
ws2["A10"] = "By type:"
for i, (t, n) in enumerate(types.most_common(), 11):
    ws2.cell(row=i, column=1, value=f"  {t}: {n}")
ws2["A1"].font = Font(bold=True, size=14)
ws2.column_dimensions["A"].width = 60

path = "output/swiss_companies.xlsx"
wb.save(path)
print(f"Saved {len(companies)} companies -> {path}")
print("By city:", dict(cities.most_common()))
print("By type:", dict(types.most_common()))