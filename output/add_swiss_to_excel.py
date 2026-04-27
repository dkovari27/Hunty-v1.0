"""
output/add_swiss_to_excel.py
Add all 113 Swiss company career pages to European_Job_Search_Websites.xlsx
as a new "Switzerland" country block.

Run once: python output/add_swiss_to_excel.py
Run from the Job-Hunter root directory.
"""
from pathlib import Path
import openpyxl

EXCEL_PATH = Path(__file__).parent.parent / "European_Job_Search_Websites.xlsx"

# (Website name, URL, Description)
SWISS_COMPANIES = [
    # ── Basel city ────────────────────────────────────────────────────────────
    ("Boehringer Ingelheim",    "https://jobs.boehringer-ingelheim.com/",               "Pharma; Basel HQ"),
    ("Regeneron",               "https://careers.regeneron.com/en/",                    "Pharma; Basel office"),
    ("Roche",                   "https://careers.roche.com/global/en",                  "Pharma; Basel HQ"),
    ("Novartis",                "https://www.novartis.com/careers",                     "Pharma; Basel HQ"),
    ("Sandoz",                  "https://www.sandoz.com/careers/",                      "Pharma generics; Basel"),
    ("Sanofi",                  "https://jobs.sanofi.com/en",                           "Pharma; Basel & Vernier"),
    ("Bayer",                   "https://www.bayer.com/en/ch/careers",                  "Pharma; Basel hub"),
    ("BeOne Medicines",         "https://beigene.wd3.myworkdayjobs.com/beigene",        "Pharma; European HQ Basel (Workday)"),
    ("Moderna",                 "https://www.modernatx.com/careers",                    "Pharma; mRNA; Basel hub"),
    ("Mepha",                   "https://www.mepha.ch/de-ch/unternehmen/karriere/",     "Pharma generics; Basel"),
    ("Pentapharm",              "https://www.pentapharm.com/careers",                   "Specialty biochemicals; Basel"),
    ("Swedish Orphan Biovitrum","https://sobi.csod.com/ats/careersite/search.aspx?site=2&c=sobi", "Pharma SOBI; Basel (Cornerstone)"),
    ("Lonza",                   "https://www.lonza.com/careers",                        "CDMO; Basel HQ"),
    ("Celonic",                 "https://celonic-ag.jobs.personio.com",                 "Biologics CDMO; Basel (Personio)"),
    ("LYO-X",                   "https://lyo-x.com/en/",                               "Lyophilisation CDMO; Basel"),
    ("ten23 health",            "https://ten23.health/careers/",                        "Drug dev CDMO; Basel"),
    ("SpiroChem",               "https://spirochem.com/careers/",                       "Discovery CRO; Basel"),
    ("Bright Peak Therapeutics","https://jobs.ashbyhq.com/brightpeak",                  "Biotech; Basel (Ashby)"),
    ("BioVersys",               "https://www.bioversys.com/careers/",                   "Biotech antibiotics; Basel"),
    ("Noema Pharma",            "https://jobs.ashbyhq.com/noema",                       "Biotech CNS; Basel (Ashby)"),
    ("Monte Rosa Therapeutics", "https://jobs.ashbyhq.com/MonteRosaTherapeutics",       "Biotech molecular glue; Basel (Ashby)"),
    ("NBE-Therapeutics",        "https://nbe-therapeutics.com/employment/vacancies/",   "Biotech ADC; Basel"),
    ("Vir Biotechnology",       "https://jobs.ashbyhq.com/vir",                         "Biotech infectious disease; Basel (Ashby)"),
    ("Roivant Sciences",        "https://jobs.ashbyhq.com/roivant",                     "Biotech; Basel (Ashby)"),
    ("FoRx Therapeutics",       "https://www.forxtherapeutics.com/career",              "Biotech; Basel"),
    ("Cimeio Therapeutics",     "https://www.cimeio.com/careers/",                      "Biotech cell & gene; Basel"),
    ("Mosanna Therapeutics",    "https://mosanna.com/careers/",                         "Biotech; Basel"),
    ("Synendos Therapeutics",   "https://www.synendos.com/open-position",               "Biotech endocannabinoid; Basel"),
    ("Tolremo Therapeutics",    "https://www.tolremo.com/careers",                      "Biotech; Basel"),
    ("Vaximm",                  "https://vaximm.com/about-us/",                         "Biotech oral vaccines; Basel"),
    ("Versameb",                "https://versameb.com/contact/career/",                 "Biotech mRNA; Basel"),
    ("Genedata",                "https://jobs.danaher.com/global/en/genedata",          "Bioinfo software; Basel (Danaher)"),
    ("Scailyte",                "https://scailyte.com/careers/",                        "Biotech single-cell AI; Basel"),
    ("Oxford BioTherapeutics",  "https://www.oxfordbiotherapeutics.com/careers",        "Biotech ADC oncology; Basel"),
    ("Skyhawk Therapeutics",    "https://job-boards.greenhouse.io/skyhawktherapeutics", "Biotech RNA splicing; Basel (Greenhouse)"),
    ("Alentis Therapeutics",    "https://jobs.ashbyhq.com/alentis",                     "Biotech fibrosis & cancer; Basel (Ashby)"),
    ("Windward Bio",            "https://jobs.ashbyhq.com/WindwardBio",                 "Biotech immunology; Basel (Ashby)"),
    ("Anaveon",                 "https://anaveon.com/careers/",                         "Biotech IL-2; Basel"),
    ("RhyGaze",                 "https://rhygaze.com/careers/",                         "Biotech gene therapy; Basel"),
    ("Captor Therapeutics",     "https://jobs.ashbyhq.com/captor",                      "Biotech molecular glue; Basel (Ashby)"),
    ("Altamira Therapeutics",   "https://www.altamiratherapeutics.com/careers/",        "Biotech RNA delivery; Basel"),
    ("Alloy Therapeutics",      "https://www.alloytx.com/careers/",                     "Biotech antibody; Basel"),
    ("Amporin Pharmaceuticals", "https://amporin.com/careers/",                         "Biotech small molecules; Basel"),
    ("AlloCyte Pharmaceuticals","https://allocyte.com/careers/",                        "Biotech allogeneic cell; Basel"),
    ("Ridgeline Discovery",     "https://ridgelinediscovery.com/careers/",              "CRO drug discovery; Basel"),
    ("WuXi XDC",                "https://www.wuxixdc.com/careers/",                     "CDMO ADC; Basel"),
    ("T3 Pharmaceuticals",      "https://www.t3pharma.com/careers/",                    "Biotech engineered bacteria; Basel"),
    ("PharmaBiome",             "https://pharmabio.me/careers/",                        "Biotech microbiome; Basel"),
    ("Granite Bio",             "https://www.granitebio.com/careers/",                  "Biotech immunology; Basel"),
    ("Stromal Therapeutics",    "https://stromaltx.com/careers/",                       "Biotech tumour stroma; Basel"),
    ("Straumann Group",         "https://www.straumann.com/group/en/discover/career.html", "MedTech dental; Basel"),
    # ── Basel suburbs ────────────────────────────────────────────────────────
    ("J&J Innovative Medicine", "https://www.careers.jnj.com/en/jobs/",                 "Pharma; Allschwil (ex-Actelion)"),
    ("Idorsia Pharmaceuticals", "https://careers.idorsia.com/",                         "Pharma; Allschwil"),
    ("Basilea Pharmaceutica",   "https://jobs.basilea.com/eng",                         "Pharma anti-infectives; Allschwil"),
    ("Spexis",                  "https://spexisbio.com/careers/",                       "Pharma rare disease; Allschwil"),
    ("Swiss Pharma Contract",   "https://swisspharmaco.com/contact/",                   "CRO/CDMO; Allschwil"),
    ("CARBOGEN AMCIS",          "https://www.carbogen-amcis.com/careers/open-positions","CDMO API & HPAPI; Bubendorf"),
    ("Bachem",                  "https://careers.bachem.com/search",                    "CDMO peptide; Bubendorf (SuccessFactors)"),
    ("Siegfried",               "https://siegfried.wd103.myworkdayjobs.com/external",   "CDMO API; Zofingen (Workday)"),
    # ── Zurich city ──────────────────────────────────────────────────────────
    ("Takeda",                  "https://jobs.takeda.com/search-jobs",                  "Pharma; Zurich HQ"),
    ("AbbVie",                  "https://careers.abbvie.com/",                          "Pharma; Zurich"),
    ("Acino",                   "https://acino.swiss/careers/",                         "Pharma specialty; Zurich"),
    ("Fosun Pharma",            "https://www.fosunpharma.com/en/careers/",              "Pharma; Zurich"),
    ("Alvotech",                "https://www.alvotech.com/about-us/careers",            "Pharma biosimilars; Zurich"),
    ("MetrioPharm",             "https://www.metriopharm.com/en/about-us/Career.html",  "Pharma anti-inflammatory; Zurich"),
    ("CDR-Life",                "https://www.cdr-life.com/careers/",                    "Biotech bispecific ab; Zurich"),
    ("Muvon Therapeutics",      "https://www.muvon-therapeutics.com/openpositions",     "Biotech cell therapy; Zurich"),
    ("Cellvie",                 "https://cellvie.bio/careers",                          "Biotech mitochondrial; Zurich"),
    ("Endogena Therapeutics",   "https://endogena.com/careers/",                        "Biotech retinal; Zurich"),
    ("Mabylon",                 "https://mabylon.com/careers/",                         "Biotech antibodies; Zurich"),
    ("SYDRA AG",                "https://sydra.io/careers/",                            "Biotech AI longevity; Zurich"),
    ("DataHow",                 "https://datahow.ch/about/careers/",                    "Biotech bioprocess AI; Zurich"),
    ("QPillars",                "https://qpillars.com/careers",                         "Pharma quality software; Zurich"),
    ("MaxWell Biosystems",      "https://www.mxwbio.com/careers",                       "Biotech MEA; Zurich"),
    ("BC Platforms",            "https://careers.bcplatforms.com/",                     "Biotech genomic platform; Zurich"),
    ("Sequana Medical",         "https://www.sequanamedical.com/open-positions/",       "MedTech fluid; Zurich"),
    ("Celerion",                "https://www.celerion.com/careers",                     "CRO clinical; Zurich"),
    # ── Zurich suburbs ───────────────────────────────────────────────────────
    ("Molecular Partners",      "https://www.molecularpartners.com/careers/",           "Biotech DARPin; Schlieren"),
    ("Philochem",               "https://www.philogen.com/en/careers",                  "Biotech drug conjugates; Otelfingen"),
    ("Tandem Therapeutics",     "https://jobs.ashbyhq.com/tandem",                      "Biotech ECM/fibrosis; Zurich (Ashby)"),
    # ── Zug / Rotkreuz ───────────────────────────────────────────────────────
    ("Amgen",                   "https://careers.amgen.com/en/Switzerland",             "Pharma; Rotkreuz"),
    ("Roche Diagnostics",       "https://careers.roche.com/global/en/rotkreuz",         "Pharma diagnostics; Rotkreuz"),
    ("AstraZeneca",             "https://careers.astrazeneca.com/location/switzerland-jobs/7684/2658434/2", "Pharma; Zug"),
    ("Biogen",                  "https://www.biogen.com/careers.html",                  "Biotech neurology; Baar"),
    ("Bristol Myers Squibb",    "https://careers.bms.com/",                             "Pharma; Steinhausen"),
    ("Gilead Sciences",         "https://jobs.gilead.com/",                             "Pharma antiviral; Zug"),
    ("GSK",                     "https://www.gsk.com/en-gb/careers/",                   "Pharma oncology unit; Zug"),
    ("Pfizer",                  "https://www.pfizer.com/people/careers",                "Pharma; Zurich/Zug"),
    ("Galderma",                "https://www.galderma.com/en/careers",                  "Pharma dermatology; Zug HQ"),
    ("Blueprint Medicines",     "https://job-boards.greenhouse.io/blueprintmedicines",  "Biotech precision oncology; Zug (Greenhouse)"),
    ("Alnylam",                 "https://www.alnylam.com/careers",                      "Biotech RNAi; Zug European HQ"),
    ("Nuvation Bio",            "https://nuvationbio.com/careers/",                     "Biotech oncology; Zug"),
    ("Apellis",                 "https://apellis.com/careers/",                         "Biotech complement; Zug"),
    ("Immunocore",              "https://jobs.ashbyhq.com/immunocore",                  "Biotech ImmTAX; Zug (Ashby)"),
    ("PSI CRO",                 "https://psicro.com/careers/",                          "CRO full-service; Zug"),
    ("PolyPeptide Group",       "https://www.polypeptide.com/careers",                  "CDMO peptide; Zug"),
    ("Recipharm",               "https://www.recipharm.com/join-us/",                   "CDMO drug product; Zug"),
    # ── Bern ─────────────────────────────────────────────────────────────────
    ("CSL Behring",             "https://www.csl.com/careers",                          "Pharma plasma-derived; Bern"),
    ("AlveoliX",                "https://www.alveolix.com/about/",                      "Biotech lung-on-chip; Bern"),
    ("CellnTec",                "https://cellntec.com/careers/",                        "Biotech epithelial cell culture; Bern"),
    ("Ypsomed",                 "https://www.ypsomed.com/en/careers.html",              "MedTech drug delivery; Burgdorf"),
    # ── Lausanne / Vaud ──────────────────────────────────────────────────────
    ("Ferring Pharmaceuticals", "https://www.ferring.com/careers/",                     "Pharma reproductive health; St-Prex"),
    ("Debiopharm",              "https://apply.workable.com/debiopharm/",               "Pharma oncology; Lausanne (Workable)"),
    ("AC Immune",               "https://www.acimmune.com/careers/",                    "Biotech neurodegeneration; Lausanne"),
    ("HAYA Therapeutics",       "https://www.hayatx.com/careers",                       "Biotech lncRNA; Lausanne"),
    ("Oculis",                  "https://oculis.com/jobs/",                             "Biotech eye disease; Lausanne"),
    ("AB2 Bio",                 "http://www.ab2bio.com/en/careers.32.html",             "Biotech IL-18; Lausanne"),
    ("Ichnos Sciences",         "https://iginnovate.com/careers/",                      "Biotech bispecific ab; Lausanne (IGI)"),
    ("Opna Bio",                "https://opnabio.com/careers/",                         "Biotech oncology; Epalinges"),
    ("Onward Therapeutics",     "https://www.onwardtherapeutics.com/careers/",          "Biotech oncology; Epalinges"),
    ("RS Research",             "https://rsresearch.net/careers/",                      "CRO clinical; Lausanne"),
    ("Novostia",                "https://www.novostia.com/careers",                     "MedTech cardiac; Lausanne"),
]


def main() -> None:
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # Check if Switzerland rows already exist (idempotent re-run)
    existing_urls = {str(row[2]).rstrip("/") for row in ws.iter_rows(min_row=2, values_only=True) if row[2]}

    added = 0
    for name, url, description in SWISS_COMPANIES:
        if url.rstrip("/") in existing_urls:
            continue
        ws.append(["Switzerland", name, url, description])
        added += 1

    wb.save(EXCEL_PATH)
    print(f"Added {added} Swiss company rows to {EXCEL_PATH}")
    print(f"Total rows now: {ws.max_row - 1}")


if __name__ == "__main__":
    main()
